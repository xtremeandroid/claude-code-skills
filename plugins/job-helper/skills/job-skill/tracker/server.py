#!/usr/bin/env python3
"""
Job application tracker web UI.

Reads the job tracker xlsx + the dated application folders (applications_YYYY-MM-DD),
matches each spreadsheet row to its generated resume/cover-letter folder, and serves a
browser UI to tick off applications, copy CV text, and bulk-open job links.

Run:  python3 tracker/server.py  [--port 8765] [--no-open]
New dated folders are picked up automatically -- nothing to configure.
"""

import argparse
import http.server
import json
import os
import re
import shutil
import socketserver
import subprocess
import sys
import threading
import time
import urllib.parse
import webbrowser
import xml.etree.ElementTree as ET
import zipfile
from datetime import date as _date

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl missing. Install with: python3 -m pip install --user openpyxl")

ROOT = os.path.dirname(os.path.abspath(os.path.dirname(__file__)))
HERE = os.path.dirname(os.path.abspath(__file__))
BACKUP_DIR = os.path.join(HERE, "backups")
FOLDER_RE = re.compile(r"^applications_(\d{4}-\d{2}-\d{2})$")
# Newer runs write to search-YYYY-MM-DD/applications/ instead of applications_YYYY-MM-DD/.
SEARCH_RE = re.compile(r"^search-(\d{4}-\d{2}-\d{2})$")
SHEET_NAME = "Applications"

# Folders inside a run that aren't a single job application — they hold one draft per
# file (a funded startup, or a recruiter post) and get expanded into their own rows.
SPECIAL = {
    "_FreshlyFunded_NoPostingYet": ("Freshly funded", "No posting yet — congrats-on-the-raise email drafted"),
    "_RecruiterPosts_ReplyDirectly": ("Recruiter post", "Reply to the person directly — draft ready"),
}

# Column header -> canonical key. Anything not listed still comes through as an extra.
COLS = {
    "Date Found": "date",
    "Company": "company",
    "Role": "role",
    "Job ID": "job_id",
    "Platform": "platform",
    "Location": "location",
    "Posted Date": "posted",
    "Job URL": "url",
    "Fitness Score": "fitness",
    "Required YoE": "yoe",
    "Salary": "salary",
    "Cover Letter": "has_cover",
    "Status": "status",
    "Date Applied": "date_applied",
    "Why It Matches": "why",
    "Next Action": "next_action",
}

_write_lock = threading.Lock()
_backed_up = set()


# ---------------------------------------------------------------- helpers

def norm(s):
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())


def lcp_len(a, b):
    n = 0
    for x, y in zip(a, b):
        if x != y:
            break
        n += 1
    return n


def dated_folders():
    """{date: abs path to the folder holding that run's per-company dirs}.

    Handles both layouts: the old applications_YYYY-MM-DD/ and the newer
    search-YYYY-MM-DD/applications/. If both exist for a date, the newer wins.
    """
    out = {}
    for name in os.listdir(ROOT):
        p = os.path.join(ROOT, name)
        if not os.path.isdir(p):
            continue
        m = FOLDER_RE.match(name)
        if m:
            out.setdefault(m.group(1), p)
            continue
        m = SEARCH_RE.match(name)
        if m:
            apps = os.path.join(p, "applications")
            if os.path.isdir(apps):
                out[m.group(1)] = apps
    return out


def run_root(base):
    """The run directory for a given applications folder (parent if it's search-*/applications)."""
    parent = os.path.dirname(base)
    return parent if SEARCH_RE.match(os.path.basename(parent) or "") else base


def outreach_index(base):
    """{normalised company: outreach record} from this run's raw/outreach.json, if present."""
    path = os.path.join(run_root(base), "raw", "outreach.json") if base else None
    if not path or not os.path.exists(path):
        return {}
    try:
        with open(path, encoding="utf-8") as fh:
            data = json.load(fh)
    except (ValueError, OSError):
        return {}
    out = {}
    for rec in data if isinstance(data, list) else []:
        key = norm(str(rec.get("company", "")).split("(")[0])
        if not key:
            continue
        contacts = [c for c in (rec.get("contacts") or []) if c.get("name")]
        out[key] = {
            "contacts": [
                {
                    "name": c.get("name") or "",
                    "title": c.get("title") or "",
                    "email": c.get("email") or "",
                    "phone": c.get("phone") or "",
                    "linkedin": c.get("linkedin") or "",
                }
                for c in contacts
            ],
            "connections": [
                {
                    "name": c.get("name") or "",
                    "degree": c.get("degree") or "",
                    "title": c.get("title") or "",
                }
                for c in (rec.get("connections") or [])
                if c.get("name")
            ],
            "generic_emails": [e for e in (rec.get("generic_emails") or []) if e],
            "best_route": rec.get("best_route") or "",
            "notes": rec.get("notes") or "",
        }
    return out


def find_workbook(date_str, folders):
    """Prefer an xlsx living inside the dated folder; fall back to a root-level tracker."""
    folder = folders.get(date_str)
    if folder:
        local = sorted(
            os.path.join(folder, f)
            for f in os.listdir(folder)
            if f.endswith(".xlsx") and not f.startswith("~$")
        )
        if local:
            return local[0]
    root_books = sorted(
        os.path.join(ROOT, f)
        for f in os.listdir(ROOT)
        if f.endswith(".xlsx") and not f.startswith("~$")
    )
    prefer = [b for b in root_books if "tracker" in os.path.basename(b).lower()]
    return (prefer or root_books or [None])[0]


def read_sheet(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], {}, ws.title
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    idx = {h: i for i, h in enumerate(header)}
    records = []
    for n, raw in enumerate(rows[1:], start=2):  # sheet row numbers are 1-based + header
        if not any(v not in (None, "") for v in raw):
            continue
        rec = {"sheet_row": n}
        for head, key in COLS.items():
            i = idx.get(head)
            v = raw[i] if i is not None and i < len(raw) else None
            rec[key] = "" if v is None else str(v).strip()
        records.append(rec)
    return records, idx, (SHEET_NAME if SHEET_NAME in wb.sheetnames else ws.title)


def list_job_folders(base):
    """Per-company folders only — the SPECIAL folders are expanded separately."""
    if not base or not os.path.isdir(base):
        return []
    return sorted(
        f for f in os.listdir(base)
        if os.path.isdir(os.path.join(base, f))
        and not f.startswith(".")
        and f not in SPECIAL
    )


def match_folders(records, base):
    """Greedy best-score assignment of spreadsheet rows to on-disk company folders."""
    folders = list_job_folders(base)
    if not folders:
        return
    scored = []
    for ri, rec in enumerate(records):
        cn = norm(rec.get("company"))
        ln = norm(rec.get("location"))
        if not cn:
            continue
        for f in folders:
            fn = norm(f)
            s = lcp_len(fn, cn)
            if s < 4 and not (s == len(cn) and s >= 3):
                continue
            if ln[:5] and ln[:5] in fn:
                s += 3
            scored.append((s, ri, f))
    scored.sort(key=lambda t: (-t[0], t[1], t[2]))
    used_r, used_f = set(), set()
    for s, ri, f in scored:
        if ri in used_r or f in used_f:
            continue
        used_r.add(ri)
        used_f.add(f)
        records[ri]["folder"] = os.path.join(base, f)
        records[ri]["folder_name"] = f
        records[ri]["docs"] = list_docs(os.path.join(base, f))


def list_docs(folder):
    docs = []
    for f in sorted(os.listdir(folder)):
        if f.startswith("~$") or f.startswith("."):
            continue
        low = f.lower()
        if "coverletter" in low or "cover" in low:
            kind = "cover"
        elif "connectionnote" in low or "connectnote" in low:
            kind = "connectnotes"
        elif "directmessage" in low:
            kind = "directmessages"
        elif "coldemail" in low:
            kind = "coldemail"
        elif "referral" in low:
            kind = "referral"
        elif low.startswith("contacts"):
            kind = "contacts"
        elif low.startswith("jobdetails"):
            kind = "details"
        elif low.startswith("reply_"):
            kind = "reply"
        elif "resume" in low or low.endswith("cv.docx"):
            kind = "resume"
        else:
            kind = "other"
        docs.append({"name": f, "kind": kind, "path": os.path.join(folder, f)})
    return docs


def docx_text(path):
    W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
    with zipfile.ZipFile(path) as z:
        xml = z.read("word/document.xml")
    root = ET.fromstring(xml)
    out = []
    for para in root.iter(W + "p"):
        out.append("".join(t.text or "" for t in para.iter(W + "t")))
    # collapse runs of blank lines
    text, blank = [], 0
    for line in out:
        if line.strip():
            blank = 0
            text.append(line)
        else:
            blank += 1
            if blank == 1:
                text.append("")
    return "\n".join(text).strip()


def doc_text(path):
    low = path.lower()
    if low.endswith(".docx"):
        return docx_text(path)
    if low.endswith((".txt", ".md")):
        with open(path, encoding="utf-8", errors="replace") as fh:
            return fh.read()
    raise ValueError("no text extractor for %s" % os.path.basename(path))


def backup_once(path):
    if path in _backed_up:
        return
    os.makedirs(BACKUP_DIR, exist_ok=True)
    stamp = time.strftime("%Y%m%d-%H%M%S")
    base = os.path.basename(path).replace(".xlsx", "")
    shutil.copy2(path, os.path.join(BACKUP_DIR, "%s_%s.xlsx" % (base, stamp)))
    _backed_up.add(path)


def set_status(path, sheet_row, applied):
    """Tick -> Applied + today's date. Untick -> Ready to Apply + cleared date."""
    with _write_lock:
        backup_once(path)
        wb = openpyxl.load_workbook(path)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.worksheets[0]
        header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        try:
            c_status = header.index("Status") + 1
        except ValueError:
            wb.close()
            raise ValueError("no Status column in sheet")
        c_date = header.index("Date Applied") + 1 if "Date Applied" in header else None
        ws.cell(row=sheet_row, column=c_status).value = "Applied" if applied else "Ready to Apply"
        if c_date:
            ws.cell(row=sheet_row, column=c_date).value = (
                _date.today().isoformat() if applied else None
            )
        wb.save(path)
        wb.close()


def open_urls(urls):
    if sys.platform == "darwin":
        cmd = ["open"]
    elif sys.platform.startswith("linux"):
        cmd = ["xdg-open"]
    else:
        for u in urls:
            webbrowser.open_new_tab(u)
        return len(urls)
    for u in urls:
        subprocess.Popen(cmd + [u], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        time.sleep(0.12)  # give the browser time to keep tab order sane
    return len(urls)


def reveal(path):
    if sys.platform == "darwin":
        subprocess.Popen(["open", "-R", path])
    elif sys.platform.startswith("linux"):
        subprocess.Popen(["xdg-open", os.path.dirname(path)])
    else:
        os.startfile(os.path.dirname(path))  # noqa: F821 (windows only)


def special_rows(base, start_id=-1000):
    """One row per draft file inside the _FreshlyFunded / _RecruiterPosts folders.

    These aren't spreadsheet jobs — there's no listing to apply to — so they get
    negative ids, a disabled tick box, and their own group in the UI.
    """
    rows = []
    if not base:
        return rows
    nid = start_id
    for folder, (label, blurb) in SPECIAL.items():
        d = os.path.join(base, folder)
        if not os.path.isdir(d):
            continue
        for f in sorted(os.listdir(d)):
            if f.startswith(".") or not f.lower().endswith((".txt", ".md")):
                continue
            stem = os.path.splitext(f)[0]
            who = re.sub(r"^(ColdEmail|Reply)_", "", stem).replace("_", " ")
            who = re.sub(r"(?<=[a-z])(?=[A-Z])", " ", who).strip()
            rows.append({
                "sheet_row": nid,
                "date": "",
                "company": who or stem,
                "role": label,
                "location": "",
                "platform": label,
                "status": "Draft ready",
                "url": "",
                "why": blurb,
                "group": label,
                "folder": d,
                "folder_name": folder,
                "docs": [x for x in list_docs(d) if x["name"] == f],
            })
            nid -= 1
    return rows


def synthetic_row(base, folder_name, sheet_row=-1):
    """A row for a folder that exists on disk but has no spreadsheet entry yet."""
    parts = folder_name.split("_")
    company = re.sub(r"(?<=[a-z])(?=[A-Z])", " ", parts[0])
    return {
        "sheet_row": sheet_row,
        "date": "",
        "company": company,
        "role": "",
        "location": " ".join(parts[1:]),
        "platform": "",
        "status": "Found",
        "url": "",
        "why": "Not in the spreadsheet yet -- listed from the folder on disk.",
        "folder": os.path.join(base, folder_name),
        "folder_name": folder_name,
        "docs": list_docs(os.path.join(base, folder_name)),
    }


def build_payload(date_str=None):
    folders = dated_folders()
    book_dates = set()
    books = {}
    for d in folders:
        b = find_workbook(d, folders)
        if b:
            books[d] = b
    root_book = find_workbook(None, folders)
    all_records = {}
    if root_book:
        recs, _, _ = read_sheet(root_book)
        for r in recs:
            if r.get("date"):
                book_dates.add(r["date"][:10])
        all_records[root_book] = recs

    dates = sorted(set(folders) | book_dates, reverse=True)
    if not dates:
        return {"dates": [], "date": None, "rows": [], "workbook": root_book}
    if date_str not in dates:
        date_str = dates[0]

    book = books.get(date_str) or root_book
    recs = all_records.get(book)
    if recs is None:
        recs, _, _ = read_sheet(book)
        all_records[book] = recs
    if book == root_book:
        recs = [dict(r) for r in recs if (r.get("date") or "")[:10] == date_str]
    else:
        recs = [dict(r) for r in recs]

    base = folders.get(date_str)
    if not recs and base:
        # Folder exists but the sheet has no rows for it yet -- still expose the documents.
        recs = [synthetic_row(base, f, -(i + 2))
                for i, f in enumerate(list_job_folders(base))]
    else:
        match_folders(recs, base)

    # Attach recruiter/founder contacts and referral routes found for this run.
    oi = outreach_index(base)
    for r in recs:
        key = norm(str(r.get("company", "")).split("(")[0])
        hit = oi.get(key)
        if not hit:
            hit = next((v for k, v in oi.items()
                        if key and (k.startswith(key[:12]) or key.startswith(k[:12]))), None)
        if hit:
            r["outreach"] = hit

    recs = recs + special_rows(base)

    for r in recs:
        r.setdefault("folder", "")
        r.setdefault("folder_name", "")
        r.setdefault("docs", [])
        r.setdefault("group", "")
        r["docs"] = [
            {"name": d["name"], "kind": d["kind"], "text": d["name"].lower().endswith((".docx", ".txt", ".md"))}
            for d in r["docs"]
        ]
    return {
        "dates": dates,
        "date": date_str,
        "rows": recs,
        "workbook": book,
        "workbook_name": os.path.basename(book) if book else "",
        "folder": folders.get(date_str, ""),
        "today": _date.today().isoformat(),
    }


def resolve_doc(date_str, sheet_row, kind, name=None):
    payload = build_payload(date_str)
    row = next((r for r in payload["rows"] if r["sheet_row"] == sheet_row), None)
    if not row or not row.get("folder"):
        raise FileNotFoundError("no documents folder for that row")
    folder = row["folder"]
    docs = list_docs(folder)
    if name:
        hit = next((d for d in docs if d["name"] == name), None)
    else:
        hit = next((d for d in docs if d["kind"] == kind), None)
    if not hit:
        raise FileNotFoundError("no %s document in %s" % (kind, os.path.basename(folder)))
    return hit


# ---------------------------------------------------------------- http

class Handler(http.server.BaseHTTPRequestHandler):
    server_version = "JobTracker/1.0"

    def log_message(self, fmt, *a):
        if os.environ.get("TRACKER_VERBOSE"):
            super().log_message(fmt, *a)

    # -- plumbing
    def _send(self, code, body, ctype="application/json; charset=utf-8"):
        if isinstance(body, (dict, list)):
            body = json.dumps(body).encode()
        elif isinstance(body, str):
            body = body.encode()
        self.send_response(code)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(body)

    def _json_body(self):
        n = int(self.headers.get("Content-Length") or 0)
        return json.loads(self.rfile.read(n) or b"{}")

    def do_GET(self):
        u = urllib.parse.urlparse(self.path)
        q = urllib.parse.parse_qs(u.query)
        try:
            if u.path in ("/", "/index.html"):
                with open(os.path.join(HERE, "index.html"), "rb") as fh:
                    return self._send(200, fh.read(), "text/html; charset=utf-8")
            if u.path == "/api/data":
                return self._send(200, build_payload((q.get("date") or [None])[0]))
            if u.path == "/api/doc":
                doc = resolve_doc(
                    (q.get("date") or [None])[0],
                    int(q["row"][0]),
                    (q.get("kind") or ["resume"])[0],
                    (q.get("name") or [None])[0],
                )
                return self._send(200, {"name": doc["name"], "text": doc_text(doc["path"])})
            return self._send(404, {"error": "not found"})
        except Exception as e:  # surface the message in the UI rather than a blank 500
            return self._send(400, {"error": str(e)})

    def do_POST(self):
        u = urllib.parse.urlparse(self.path)
        try:
            body = self._json_body()
            if u.path == "/api/status":
                payload = build_payload(body.get("date"))
                if not payload.get("workbook"):
                    raise ValueError("no workbook found")
                sheet_row = int(body["row"])
                if sheet_row < 2:
                    raise ValueError("this row is not in the spreadsheet yet")
                set_status(payload["workbook"], sheet_row, bool(body["applied"]))
                return self._send(200, {"ok": True})
            if u.path == "/api/open":
                urls = [u2 for u2 in body.get("urls", []) if u2.startswith(("http://", "https://"))]
                return self._send(200, {"opened": open_urls(urls[:40])})
            if u.path == "/api/reveal":
                p = body.get("path", "")
                if not os.path.abspath(p).startswith(ROOT) or not os.path.exists(p):
                    raise ValueError("bad path")
                reveal(p)
                return self._send(200, {"ok": True})
            if u.path == "/api/open-doc":
                doc = resolve_doc(body.get("date"), int(body["row"]), body.get("kind", "resume"), body.get("name"))
                subprocess.Popen(["open", doc["path"]] if sys.platform == "darwin" else ["xdg-open", doc["path"]])
                return self._send(200, {"ok": True})
            return self._send(404, {"error": "not found"})
        except Exception as e:
            return self._send(400, {"error": str(e)})


class Server(socketserver.ThreadingTCPServer):
    allow_reuse_address = True
    daemon_threads = True


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--port", type=int, default=8765)
    ap.add_argument("--no-open", action="store_true")
    args = ap.parse_args()

    port = args.port
    for _ in range(20):
        try:
            httpd = Server(("127.0.0.1", port), Handler)
            break
        except OSError:
            port += 1
    else:
        sys.exit("no free port near %d" % args.port)

    url = "http://127.0.0.1:%d/" % port
    print("job tracker  ->  %s" % url)
    print("watching     :  %s" % ROOT)
    if not args.no_open:
        threading.Timer(0.4, lambda: webbrowser.open_new_tab(url)).start()
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        print("\nbye")


if __name__ == "__main__":
    main()
